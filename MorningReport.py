import sys
import os
import django
import datetime
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
from Variables import *

#*****************************
sys.path.append(project_location)#folder for the app
#*****************************
os.environ['DJANGO_SETTINGS_MODULE'] = 'testwebpage.settings'
django.setup()

from prodfloor.models import Info,Times,Stops
from extra_functions import *

#Change this values if first setup
#*****************************
u_input = input('Please insert the date(mm/dd/yyyy):')
dates_list = u_input.split('/')
user = user_var
day = datetime.date(int(dates_list[2]),int(dates_list[0]),int(dates_list[1]))
#*****************************

grayFill = PatternFill(start_color='CBFCC4',end_color = 'CBFCC4',fill_type = 'solid')
date_style = NamedStyle(name='date_time')
date_style.number_format = 'mm/dd/yyyy'
wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Morning_Report_Template.xlsx')
minitab_wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Morning_Report_Template_minitab.xlsx')
data_wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Data.xlsx')
stop_wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Stops_Report_Template.xlsx')
sheet = wb.get_sheet_by_name('Jobs_Completed')
minitab_sheet = minitab_wb.get_sheet_by_name('Jobs_Completed')
minitab_grouped_sheet = minitab_wb.get_sheet_by_name('Jobs_Completed_Grouped')
stop_sheet = stop_wb.get_sheet_by_name('Report')
data_sheet = data_wb.get_sheet_by_name('Data')
job = Info.objects.all()
completed_before = datetime.datetime.combine(day, datetime.datetime.max.time())
completed_after = datetime.datetime.combine(day, datetime.datetime.min.time())
times = Times.objects.filter(end_time_4__gte=completed_after,end_time_4__lte=completed_before)
stops = Stops.objects.filter(stop_end_time__gte=completed_after,stop_end_time__lte=completed_before)
times_pks=[]
for i in times:
    times_pks.append(i.info.pk)
job = job.filter(pk__in=times_pks).filter(status='Complete')
completed_pks = []
for i in job:
    completed_pks.append(i.po)
completed_jobs_number = len(completed_pks)
completed_jobs = Info.objects.filter(po__in=completed_pks).order_by('po')#might not be needed
completed_jobs_m4000 = job.filter(job_type='4000')
completed_jobs_m2000 = job.filter(job_type='2000')
completed_jobs_elem = job.filter(job_type='ELEM')
c = 2
mt_c = 2
mtg_c = 2
stations_dict = {'0': '-----',
                     '1': 'S1',
                     '2': 'S2',
                     '3': 'S3',
                     '4': 'S4',
                     '5': 'S5',
                     '6': 'S6',
                     '7': 'S7',
                     '8': 'S8',
                     '9': 'S9',
                     '10': 'S10',
                     '11': 'S11',
                     '12': 'S12',
                     '13': 'ELEM1',
                     '14': 'ELEM2'}
job_type_dict = {'2000': 'M2000',
                     '4000': 'M4000',
                     'ELEM': 'Element'}
for obj_main in job:
    all_jobs = Info.objects.filter(po=obj_main.po).order_by('po')
    total_elapsed_time = datetime.timedelta(0)
    total_stops = 0
    total_time_on_stop = datetime.timedelta(0)
    total_eff_time = datetime.timedelta(0)
    for obj in all_jobs:
        start = gettimes(obj.pk, 'start')
        end = gettimes(obj.pk, 'end')
        beginning_time = spentTime(obj.pk, 1)
        program_time = spentTime(obj.pk, 2)
        logic_time = spentTime(obj.pk, 3)
        ending_time = spentTime(obj.pk, 4)
        number_of_stops = stopsnumber(obj.pk)
        time_on_stop = timeonstop(obj.pk)
        elapsed_time = totaltime(obj.pk)
        eff_time = effectivetime(obj.pk)
        tech = gettech(obj.pk)
        category = categories(obj.pk)
        total_stops += number_of_stops
        total_elapsed_time += elapsed_time
        total_time_on_stop += time_on_stop
        total_eff_time += eff_time
        sheet['A' + str(c)] = obj.job_num
        sheet['B' + str(c)] = obj.po
        sheet['C' + str(c)] = job_type_dict[obj.job_type]
        sheet['D' + str(c)] = obj.status
        sheet['E' + str(c)] = stations_dict[obj.station]
        sheet['F' + str(c)] = datetime.datetime.strptime(start,"%m/%d/%Y").date()
        sheet['F' + str(c)].style = 'DATE'
        if end != '-' and end != 'N/A':
            sheet['G' + str(c)] = datetime.datetime.strptime(end,"%m/%d/%Y").date()
        else:
            sheet['G' + str(c)] = '-'
        sheet['G' + str(c)].style = 'DATE'
        sheet['H' + str(c)] = (beginning_time.total_seconds()) / 3600
        sheet['I' + str(c)] = (program_time.total_seconds()) / 3600
        sheet['J' + str(c)] = (logic_time.total_seconds()) / 3600
        sheet['K' + str(c)] = (ending_time.total_seconds()) / 3600
        sheet['L' + str(c)] = (elapsed_time.total_seconds()) / 3600
        sheet['M' + str(c)] = number_of_stops
        sheet['N' + str(c)] = (time_on_stop.total_seconds()) / 3600
        sheet['O' + str(c)] = (eff_time.total_seconds()) / 3600
        sheet['P' + str(c)] = category
        sheet['Q' + str(c)] = tech
        minitab_sheet['A' + str(mt_c)] = obj.job_num
        minitab_sheet['B' + str(mt_c)] = obj.po
        minitab_sheet['C' + str(mt_c)] = job_type_dict[obj.job_type]
        minitab_sheet['D' + str(mt_c)] = obj.status
        minitab_sheet['E' + str(mt_c)] = stations_dict[obj.station]
        minitab_sheet['F' + str(mt_c)] = datetime.datetime.strptime(start,"%m/%d/%Y").date()
        minitab_sheet['F' + str(mt_c)].style = 'DATE1'
        if end != '-' and end != 'N/A':
            minitab_sheet['G' + str(mt_c)] = datetime.datetime.strptime(end,"%m/%d/%Y").date()
        else:
            minitab_sheet['G' + str(mt_c)] = '-'
        minitab_sheet['G' + str(mt_c)].style = 'DATE1'
        minitab_sheet['H' + str(mt_c)] = (beginning_time.total_seconds())/3600
        minitab_sheet['I' + str(mt_c)] = (program_time.total_seconds())/3600
        minitab_sheet['J' + str(mt_c)] = (logic_time.total_seconds())/3600
        minitab_sheet['K' + str(mt_c)] = (ending_time.total_seconds())/3600
        minitab_sheet['L' + str(mt_c)] = (elapsed_time.total_seconds())/3600
        minitab_sheet['M' + str(mt_c)] = number_of_stops
        minitab_sheet['N' + str(mt_c)] = (time_on_stop.total_seconds())/3600
        minitab_sheet['O' + str(mt_c)] = (eff_time.total_seconds())/3600
        minitab_sheet['P' + str(mt_c)] = category
        minitab_sheet['Q' + str(mt_c)] = tech
        c += 1
        mt_c += 1
    sheet['A' + str(c)] = obj_main.job_num
    sheet['B' + str(c)] = obj_main.po
    sheet['C' + str(c)] = job_type_dict[obj_main.job_type]
    sheet['D' + str(c)] = obj_main.status
    sheet['E' + str(c)] = '-'
    sheet['F' + str(c)] = '-'
    sheet['G' + str(c)] = '-'
    sheet['H' + str(c)] = '-'
    sheet['I' + str(c)] = '-'
    sheet['J' + str(c)] = '-'
    sheet['K' + str(c)] = '-'
    sheet['L' + str(c)] = (total_elapsed_time.total_seconds())/3600
    sheet['M' + str(c)] = total_stops
    sheet['N' + str(c)] = (total_time_on_stop.total_seconds())/3600
    sheet['O' + str(c)] = (total_eff_time.total_seconds())/3600
    sheet['P' + str(c)] = categories(obj_main.pk)
    sheet['Q' + str(c)] = '-'
    #To minitab grouped jobs sheet
    minitab_grouped_sheet['A' + str(mtg_c)] = obj_main.job_num
    minitab_grouped_sheet['B' + str(mtg_c)] = obj_main.po
    minitab_grouped_sheet['C' + str(mtg_c)] = job_type_dict[obj_main.job_type]
    minitab_grouped_sheet['D' + str(mtg_c)] = obj_main.status
    minitab_grouped_sheet['E' + str(mtg_c)] = '-'
    minitab_grouped_sheet['F' + str(mtg_c)] = '-'
    minitab_grouped_sheet['G' + str(mtg_c)] = '-'
    minitab_grouped_sheet['H' + str(mtg_c)] = '-'
    minitab_grouped_sheet['I' + str(mtg_c)] = '-'
    minitab_grouped_sheet['J' + str(mtg_c)] = '-'
    minitab_grouped_sheet['K' + str(mtg_c)] = '-'
    minitab_grouped_sheet['L' + str(mtg_c)] = (total_elapsed_time.total_seconds())/3600
    minitab_grouped_sheet['M' + str(mtg_c)] = total_stops
    minitab_grouped_sheet['N' + str(mtg_c)] = (total_time_on_stop.total_seconds())/3600
    minitab_grouped_sheet['O' + str(mtg_c)] = (total_eff_time.total_seconds())/3600
    minitab_grouped_sheet['P' + str(mtg_c)] = categories(obj_main.pk)
    minitab_grouped_sheet['Q' + str(mtg_c)] = '-'
    for rowOfCellObjects in sheet['A' + str(c):'Q' + str(c)]:
        for cellObj in rowOfCellObjects:
            cellObj.fill = grayFill
    c += 1
    mtg_c += 1

efficiency_2000 = efficiency(completed_jobs_m2000)
efficiency_4000 = efficiency(completed_jobs_m4000)
efficiency_elem = efficiency(completed_jobs_elem)
data_sheet['C3'] = completed_jobs_m4000.count()
data_sheet['C6'] = completed_jobs_m2000.count()
data_sheet['C9'] = completed_jobs_elem.count()
data_sheet['B3'] = efficiency_4000
data_sheet['B6'] = efficiency_2000
data_sheet['B9'] = efficiency_elem
data_sheet['J1'] = day.strftime("%m/%d/%Y")
data_sheet['J1'].style = 'Date'
stop_c = 2
for stop in stops:
    if stop.stop_end_time > stop.stop_start_time:
        job_obj = Info.objects.get(pk=stop.info_id)
        tech = gettech(job_obj.pk)
        stop_sheet['A' + str(stop_c)] = job_obj.job_num
        stop_sheet['B' + str(stop_c)] = job_obj.po
        stop_sheet['C' + str(stop_c)] = job_type_dict[job_obj.job_type]
        stop_sheet['D' + str(stop_c)] = stop.stop_start_time
        stop_sheet['E' + str(stop_c)] = stop.stop_end_time
        stop_sheet['F' + str(stop_c)] = stop.reason
        stop_sheet['G' + str(stop_c)] = stop.extra_cause_1
        stop_sheet['H' + str(stop_c)] = stop.extra_cause_2
        stop_sheet['I' + str(stop_c)] = stop.reason_description
        stop_sheet['J' + str(stop_c)] = stop.solution
        stop_sheet['K' + str(stop_c)] = job_obj.station
        stop_sheet['L' + str(stop_c)] = timeonstop_1(stop.pk)
        stop_sheet['M' + str(stop_c)] = tech
        stop_c += 1

wb.save(r'C:\Users\\' + user + '\Desktop\Morning Reports\Morning_Report_'+ str(day) + '.xlsx')
data_wb.save(r'C:\Users\\' + user + '\Desktop\Morning Reports\Data.xlsx')
minitab_wb.save(r'C:\Users\\' + user + '\Desktop\Morning Reports\Morning_Report_'+ str(day) + '_mt.xlsx')
stop_wb.save(r'C:\Users\\' + user + '\Desktop\Morning Reports\Stop_Report_'+ str(day) + '.xlsx')
