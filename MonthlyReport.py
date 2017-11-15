import sys
import os
import django
import datetime
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
from Variables import *

#*****************************
sys.path.append(project_location)#folder for the app
#*****************************
os.environ['DJANGO_SETTINGS_MODULE'] = 'testwebpage.settings'
django.setup()

from prodfloor.models import Info,Times,Stops
from extra_functions import *

#Change this values if first setup
#*****************************
u_input_start = input('Please insert the start date(mm/dd/yyyy):')
u_input_end = input('Please insert the end date(mm/dd/yyyy):')
start_dates_list = u_input_start.split('/')
end_dates_list = u_input_end.split('/')
user = user_var
start_day = datetime.date(int(start_dates_list[2]),int(start_dates_list[0]),int(start_dates_list[1]))
end_day = datetime.date(int(end_dates_list[2]),int(end_dates_list[0]),int(end_dates_list[1]))
#*****************************

grayFill = PatternFill(start_color='CBFCC4',end_color = 'CBFCC4',fill_type = 'solid')
stop_wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Stops_Report_Template_New.xlsx')
stop_sheet = stop_wb.get_sheet_by_name('Report')
completed_jobs_sheet = stop_wb.get_sheet_by_name('Completed_Jobs')
job = Info.objects.all()
completed_before = datetime.datetime.combine(end_day, datetime.datetime.max.time())
completed_after = datetime.datetime.combine(start_day, datetime.datetime.min.time())
stops = Stops.objects.filter(stop_end_time__gte=completed_after,stop_end_time__lte=completed_before)
times_completed_in_period = Times.objects.filter(end_time_4__gte=completed_after,end_time_4__lte=completed_before)

weeks_to_get = (end_day-start_day).days//7

c = 2
mt_c = 2
mtg_c = 2
stations_dict = {'0': '-----',
                     '1': 'S1',
                     '2': 'S2',
                     '3': 'S3',
                     '4': 'S4',
                     '5': 'S5',
                     '6': 'S6',
                     '7': 'S7',
                     '8': 'S8',
                     '9': 'S9',
                     '10': 'S10',
                     '11': 'S11',
                     '12': 'S12',
                     '13': 'ELEM1',
                     '14': 'ELEM2'}
job_type_dict = {'2000': 'M2000',
                     '4000': 'M4000',
                     'ELEM': 'Element'}
stop_c = 2

for stop in stops:
    if stop.stop_end_time > stop.stop_start_time:
        job_obj = Info.objects.get(pk=stop.info_id)
        tech = gettech(job_obj.pk)
        days = (stop.stop_end_time.date() - start_day).days#
        stop_sheet['A' + str(stop_c)] = job_obj.job_num
        stop_sheet['B' + str(stop_c)] = job_obj.po
        stop_sheet['C' + str(stop_c)] = job_type_dict[job_obj.job_type]
        stop_sheet['D' + str(stop_c)] = (days//7)+1
        stop_sheet['E' + str(stop_c)] = datetime.datetime.__format__(stop.stop_start_time.astimezone(instance_time_zone),"%m/%d/%Y %H:%M:%S")
        stop_sheet['F' + str(stop_c)] = datetime.datetime.__format__(stop.stop_end_time.astimezone(instance_time_zone),"%m/%d/%Y %H:%M:%S")
        stop_sheet['G' + str(stop_c)] = stop.reason
        stop_sheet['H' + str(stop_c)] = stop.extra_cause_1
        stop_sheet['I' + str(stop_c)] = stop.extra_cause_2
        stop_sheet['J' + str(stop_c)] = stop.reason_description
        stop_sheet['K' + str(stop_c)] = stop.solution
        stop_sheet['L' + str(stop_c)] = job_obj.station
        stop_sheet['M' + str(stop_c)] = timeonstop_1(stop.pk)
        stop_sheet['N' + str(stop_c)] = tech
        stop_c += 1


weeks_count = 0
last_day = copy.deepcopy(start_day)
jobs_per_week = {}

while weeks_count <= weeks_to_get:
    times_per_week = times_completed_in_period.filter(end_time_4__gte=last_day,end_time_4__lte=last_day + datetime.timedelta(days=7))
    completed_jobs_per_week = []
    for time_obj in times_per_week:
        completed_jobs_per_week.append(time_obj.info.po)
    jobs_per_week_per_type = {}
    jobs_completed = Info.objects.filter(po__in=completed_jobs_per_week, status="Complete")
    jobs_completed_elem = jobs_completed.filter(job_type="ELEM").count()
    jobs_completed_2000 = jobs_completed.filter(job_type="2000").count()
    jobs_completed_4000 = jobs_completed.filter(job_type="4000").count()
    jobs_per_week_per_type["Element"] = jobs_completed_elem
    jobs_per_week_per_type["M2000"] = jobs_completed_2000
    jobs_per_week_per_type["M4000"] = jobs_completed_4000
    jobs_per_week[weeks_count] = jobs_per_week_per_type

    last_day += datetime.timedelta(days=7)
    weeks_count += 1

row = 2
for week_key in jobs_per_week:
    for job_type_key in jobs_per_week[week_key]:
        completed_jobs_sheet['A' + str(row)] = job_type_key
        completed_jobs_sheet['B' + str(row)] = week_key + 1
        completed_jobs_sheet['C' + str(row)] = jobs_per_week[week_key][job_type_key]
        row += 1

stop_wb.save(r'C:\Users\\' + user + '\Desktop\Morning Reports\Stop_Report_'+ str(start_day) + '_' + str(end_day) + '.xlsx')
