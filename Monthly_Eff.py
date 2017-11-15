import sys
import os
import django
import datetime
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
from Variables import *
#*****************************
sys.path.append(project_location)#folder for the app
#*****************************
os.environ['DJANGO_SETTINGS_MODULE'] = 'testwebpage.settings'
django.setup()

from prodfloor.models import Info,Times,Stops
from prodfloor.dicts import percentage_of_time, times_per_category
from extra_functions import *
from django.contrib.auth.models import User

#Change this values if first setup
#*****************************
u_input_start = input('Please insert the start date(mm/dd/yyyy):')
u_input_end = input('Please insert the end date(mm/dd/yyyy):')
start_dates_list = u_input_start.split('/')
end_dates_list = u_input_end.split('/')
user = user_var
start_day = datetime.date(int(start_dates_list[2]),int(start_dates_list[0]),int(start_dates_list[1]))
end_day = datetime.date(int(end_dates_list[2]),int(end_dates_list[0]),int(end_dates_list[1]))
#*****************************
'''
Descripcion de este reporte:
Sacar jobs trabajados per Tech, y obtener la eficiencia promedio.
Posibles pasos para lograrlo,
    funcion que regresa todos los trabajos que hicieron los techs en el periodo de tiempo(jobs_per_time):
        toma como el parametro periodo de tiempo regresa un diccionario con el siguiente formato:
                                                        dict = {tech_id1 = {po1:[percentage,Efficiency],
                                                                            po2:[percentage,Efficiency]},
                                                                tech_id2 = {po:[percentage,Efficiency]},
                                                                tech_id3 = {po:[percentage,Efficiency]},}
        en caso de que un trabajo le haya sido reasignado multiples veces se sumara el porcentaje y 
        se promediara la eficiencia

    funcion que toma el diccionario y pasa la informacion a excel(to_excel):
        toma como parametros el diccionario generado en jobs_per_time, grabara todos los trabajos que hizo en excel, y
        los que el tecnico hay trabajo mas de 60% se promediaran y se determinara la eficiencia promedio en el periodo
        
        
Preguntas//A definir:
    Cuando multiples tecnicos trabajan en la misma etapa; como definir tiempo/porcentaje trabajado de ella, opciones:
        -Sumar las horas efectivas y determinar cuanto porcentaje le corresponde
        -El que termina la etapa se toma como el que la hizo
'''


eff_wb = openpyxl.load_workbook(r'C:\Users\\' + user + '\Desktop\Morning Reports\Efficiency_Template.xlsx')
eff_sheet = eff_wb.get_sheet_by_name('Efficiency')
completed_before = datetime.datetime.combine(end_day, datetime.datetime.max.time())
completed_after = datetime.datetime.combine(start_day, datetime.datetime.min.time())
times = Times.objects.filter(end_time_4__gte=completed_after,end_time_4__lte=completed_before)
times_pks = []
# get a list with the pk's of the jobs which end date is between the wanted dates
for i in times:
    times_pks.append(i.info.pk)
job = Info.objects.all().filter(pk__in=times_pks).filter(status='Complete')
completed_pks = []
# get a list with the po's of the jobs which end date is between the wanted dates and with the status == 'Complete'
for i in job:
    completed_pks.append(i.po)
completed_jobs_number = len(completed_pks)
completed_jobs = Info.objects.filter(po__in=completed_pks).order_by('po')  # objs finished

# function to compare an obj against all the objs with matching po, this is to determine the percentage done by ea tech
# when multiple techs
def compare_with_all_po(pk):
    job_to_compare = Info.objects.get(pk=pk)
    jobs = Info.objects.filter(po=job_to_compare.po)
    job_to_compare_category = categories(job_to_compare.pk)
    times_dict = {'Beginning': datetime.timedelta(0),
                  'Program': datetime.timedelta(0),
                  'Logic': datetime.timedelta(0),
                  'Ending': datetime.timedelta(0)}
    times_dict_to_compare = {'Beginning': datetime.timedelta(0),
                             'Program': datetime.timedelta(0),
                             'Logic': datetime.timedelta(0),
                             'Ending': datetime.timedelta(0)}
    for job_obj in jobs:
        times_dict['Beginning'] += time_per_stage(job_obj.pk,1)
        times_dict['Program'] += time_per_stage(job_obj.pk,2)
        times_dict['Logic'] += time_per_stage(job_obj.pk,3)
        times_dict['Ending'] += time_per_stage(job_obj.pk,4)
    times_dict_to_compare['Beginning'] += time_per_stage(job_to_compare.pk, 1)
    times_dict_to_compare['Program'] += time_per_stage(job_to_compare.pk, 2)
    times_dict_to_compare['Logic'] += time_per_stage(job_to_compare.pk, 3)
    times_dict_to_compare['Ending'] += time_per_stage(job_to_compare.pk, 4)
    percentage_dict = {'Beginning': 0,
                       'Program': 0,
                       'Logic': 0,
                       'Ending': 0}
    percentage_dict['Beginning'] = times_dict_to_compare['Beginning'] / times_dict['Beginning']
    percentage_dict['Program'] = times_dict_to_compare['Program'] / times_dict['Program']
    if job_to_compare.job_type != 'ELEM':
        percentage_dict['Logic'] = times_dict_to_compare['Logic'] / times_dict['Logic']
    percentage_dict['Ending'] = times_dict_to_compare['Ending'] / times_dict['Ending']
    percentage_per_category = percentage_of_time[job_to_compare.job_type]
    percentage_dict['Beginning'] = percentage_dict['Beginning'] * percentage_per_category['Beginning']
    percentage_dict['Program'] = percentage_dict['Program'] * percentage_per_category['Program']
    if job_to_compare.job_type != 'ELEM':
        percentage_dict['Logic'] = percentage_dict['Logic'] * percentage_per_category['Logic']
    percentage_dict['Ending'] = percentage_dict['Ending'] * percentage_per_category['Ending']
    percentage_of_job_done = sum(percentage_dict.values())
    time_per_category = times_per_category[job_to_compare.job_type][job_to_compare_category]
    expected_time = (percentage_of_job_done * time_per_category)/60  # time in hours
    elapsed_time = (effectivetime(job_to_compare.pk).total_seconds())/3600
    if elapsed_time <= 0 or expected_time <= 0:
        eff = 0
    else:
        eff = (expected_time/elapsed_time) * 100  # percentage
    return {'percentage': percentage_of_job_done, 'efficiency': eff, 'elapsed_time': elapsed_time,
            'expected_time': expected_time}

# function which will return a dictionary divided by techs which will contain a dict with the jobs in which the tech
# worked during the dates
def jobs_per_time(completed_jobs_objs):
    techs = User.objects.filter(groups__name='Technicians')
    dict_to_return ={}
    for tech in techs:
        jobs_per_tech = completed_jobs_objs.filter(Tech_name=tech.get_full_name())
        dict_to_return[tech.id] = {}
        for job_obj in jobs_per_tech:
            returned_dict = compare_with_all_po(job_obj.pk)
            percentage = returned_dict['percentage']
            efficiency_per_job = returned_dict['efficiency']
            elapsed_time = returned_dict['elapsed_time']
            expected_time = returned_dict['expected_time']
            if job_obj.po in dict_to_return[tech.id]:
                dict_to_return[tech.id][job_obj.po][0] += percentage
                dict_to_return[tech.id][job_obj.po][1].append(efficiency_per_job)
                dict_to_return[tech.id][job_obj.po][2] += elapsed_time
                dict_to_return[tech.id][job_obj.po][3].append(job_obj.id)
                dict_to_return[tech.id][job_obj.po][4] += expected_time
            else:
                dict_to_return[tech.id][job_obj.po] = [percentage, [efficiency_per_job], elapsed_time, [job_obj.id], expected_time]
        print(dict_to_return[tech.id])
    return dict_to_return

# function to make the writing to excel more obj oriented
def write_to_excel(col,cell_num,value,style):
    eff_sheet[col + str(cell_num)] = value
    eff_sheet[col + str(cell_num)].style = style

# might not be required
def fefficieny(eff_list):
    print(eff_list)
    count = len(eff_list)
    i_sum = sum(eff_list)
    value = i_sum/count
    print(value)
    return value

# group the stations and number of stops for jobs with multiple reassigns to the same tech
def mu_ids(ids_list):
    mu_dict = {}
    job_0 = Info.objects.get(pk=ids_list[0])
    mu_dict['stations'] = str(job_0.station)
    mu_dict['nu_stops'] = stopsnumber(job_0.pk)
    count = 1
    while count < len(ids_list):
        job_1 = Info.objects.get(pk=ids_list[count])
        mu_dict['stations'] += ', ' + str(job_1.station)
        mu_dict['nu_stops'] += stopsnumber(job_1.pk)
        count += 1
    return mu_dict

# function to determine the average efficiency
def get_av_eff(techs_dict):
    eff_sum = 0
    count = 0
    for j in techs_dict:
        if techs_dict[j][0] > 0.6:
            count += 1
            eff_sum += techs_dict[j][0]
    if eff_sum == 0 or count == 0:
        eff_av = 0
    else:
        eff_av = eff_sum/count
    return round(eff_av*100,2)

# function to pass the dictionary generated in jobs_per_time to excel workbook
def dict_to_excel(jobs_done_in_time):
    eff_sheet['B2'] = datetime.datetime.strptime(datetime.date.__format__(start_day,"%m/%d/%Y"),"%m/%d/%Y").date()
    eff_sheet['D2'] = datetime.datetime.strptime(datetime.date.__format__(end_day,"%m/%d/%Y"),"%m/%d/%Y").date()
    eff_sheet['B2'].style = 'DATE'
    eff_sheet['D2'].style = 'DATE'
    cell = 4
    for k in jobs_done_in_time:
        tech = User.objects.get(pk=k)
        jobs_dict = jobs_done_in_time[k]
        write_to_excel('A',cell,'Tech:','Check Cell')
        eff_sheet.merge_cells('B' + str(cell)+':'+'C' + str(cell))
        write_to_excel('B', cell, tech.get_full_name(),'Check Cell')
        eff_sheet.merge_cells('F' + str(cell) + ':' + 'G' + str(cell))
        write_to_excel('F', cell, 'Average Eff:','Check Cell')
        write_to_excel('H', cell, str(get_av_eff(jobs_done_in_time[k]))+'%','Check Cell')
        for rowOfCellObjects in eff_sheet['B' + str(cell):'G' + str(cell)]:
            for cellObj in rowOfCellObjects:
                cellObj.style = 'Check Cell'
        cell += 1
        write_to_excel('A', cell, 'Job No.','Check Cell')
        write_to_excel('B', cell, 'PO','Check Cell')
        write_to_excel('C', cell, 'Percentage','Check Cell')
        write_to_excel('D', cell, 'Elapsed Time','Check Cell')
        write_to_excel('E', cell, 'Efficiency','Check Cell')
        write_to_excel('F', cell, 'Station','Check Cell')
        write_to_excel('G', cell, 'Shift','Check Cell')
        write_to_excel('H', cell, 'No. Stops','Check Cell')
        cell += 1
        for po in jobs_dict:
            dict_for_po = mu_ids(jobs_dict[po][3])
            station = dict_for_po['stations']
            num_stops = dict_for_po['nu_stops']
            job_obj = Info.objects.get(pk=jobs_dict[po][3][0])
            average_efficiency_to_be_passed = fefficieny(jobs_dict[po][1])
            write_to_excel('A', cell, job_obj.job_num, 'Borders')
            write_to_excel('B', cell, po, 'Borders')
            write_to_excel('C', cell, jobs_dict[po][0]*100, 'Borders')
            write_to_excel('D', cell, jobs_dict[po][2], 'Borders')
            write_to_excel('E', cell, (jobs_dict[po][2]/jobs_dict[po][4])*100, 'Borders')
            write_to_excel('F', cell, station, 'Borders')
            write_to_excel('G', cell, 'Pending', 'Borders')
            write_to_excel('H', cell, num_stops, 'Borders')
            cell += 1
        cell += 2

dict_to_pass = jobs_per_time(completed_jobs)
dict_to_excel(dict_to_pass)


eff_wb.save(r'C:\Users\\'+user+'\Desktop\Morning Reports\Efficiency_Report_'+str(start_day)+'to'+str(end_day)+'.xlsx')
